from flask import Blueprint, request, jsonify, make_response, send_file
import io
from flask_jwt_extended import jwt_required
from bson import ObjectId
from pymongo import UpdateOne
from werkzeug.security import generate_password_hash
from datetime import datetime
import pytz

from .. import mongo
from ..utils.auth import admin_required, get_current_user
from ..utils.time_utils import (
    is_voting_window_open, format_ist, now_ist, IST, suggested_window_for_slot
)
from ..utils.export import build_csv_report
from ..services.weather import get_forecast_for_slot
from ..utils.passwords import validate_password, generate_temp_password

admin_bp = Blueprint("admin", __name__)


def _user_to_dict(u):
    # Every captain is also a player — they should be able to cast their own
    # vote even on a weekend their team has no match scheduled — so this isn't
    # a one-off flag, it's implied by the role.
    is_player = True if u["role"] == "captain" else u.get("is_player", False)
    return {
        "id": str(u["_id"]),
        "name": u["name"],
        "team_code": u["team_code"],
        "team_name": u.get("team_name", ""),
        "role": u["role"],
        "is_active": u.get("is_active", True),
        "is_player": is_player,
        "is_admin": u.get("is_admin", False),
        "matches_scheduled": u.get("matches_scheduled", 0),
        "matches_played": u.get("matches_played", 0),
        "tournament_status": u.get("tournament_status", "not_played"),
        "auction_category": u.get("auction_category"),
        "attendance_count": u.get("attendance_count", 0),
        "knockout_eligible": u.get("knockout_eligible", False),
        "batting_average": u.get("batting_average"),
        "bowling_average": u.get("bowling_average"),
        "must_change_password": u.get("must_change_password", False),
        "device_locked": bool(u.get("device_id")),
        "created_at": format_ist(u.get("created_at")),
    }


AUCTION_CATEGORIES = {"extra_power_allrounder", "extra_power_batsman", "power", "classic"}

# Normally captain/player role IS the voter roster. A handful of accounts
# (role=="admin") are also flagged is_player=True so the same admin login can
# vote too, without becoming a captain or player account — this OR-arm is what
# lets those accounts show up everywhere a voter roster is built (dashboard,
# summary, exports, auction pool).
VOTER_FILTER = {"$or": [{"role": {"$in": ["captain", "player"]}}, {"is_player": True}]}

# Narrower than VOTER_FILTER — used only by the player-management mutation
# routes below, so they additionally match admin+is_player accounts without
# also picking up captains who happen to have is_player=True (those already
# have their own update_captain/remove_captain routes; letting update_player
# match them too would just create two overlapping edit paths for the same row).
ADMIN_VOTER_FILTER = {"role": "admin", "is_player": True}


def _reset_password(user_id, role):
    """Shared by the captain/player reset-password routes below. Generates a
    fresh temp password (never needs the old one), forces a change on next
    login, revokes any session already in that account's hands (same
    token_version mechanism as self-service change-password), and logs who
    did it to whom — the accountability trail requirement #3.4 asked for."""
    acting_admin = get_current_user()
    target = mongo.db.users.find_one({"_id": ObjectId(user_id), "role": role})
    if not target:
        return jsonify({"error": f"{role.title()} not found"}), 404

    temp_password = generate_temp_password()
    mongo.db.users.update_one(
        {"_id": target["_id"]},
        {"$set": {"password_hash": generate_password_hash(temp_password), "must_change_password": True},
         "$inc": {"token_version": 1}},
    )
    mongo.db.password_resets.insert_one({
        "admin_id": str(acting_admin["_id"]),
        "admin_name": acting_admin["name"],
        "target_user_id": str(target["_id"]),
        "target_user_name": target["name"],
        "target_role": role,
        "reset_at": datetime.utcnow(),
    })
    return jsonify({"message": "Password reset", "temp_password": temp_password})


def _parse_average(value):
    """None clears the stat (not yet recorded); anything else must be a
    non-negative number — a captain's release order shouldn't silently break
    on a bad value slipping through."""
    if value is None or value == "":
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        raise ValueError("must be a number")
    if parsed < 0:
        raise ValueError("must not be negative")
    return parsed


def _slot_to_dict(slot):
    return {
        "id": str(slot["_id"]),
        "slot_number": slot["slot_number"],
        "day": slot["day"],
        "time_of_day": slot["time_of_day"],
        "match_time": slot.get("match_time", ""),
        "match_date": slot.get("match_date"),
        "description": slot.get("description", ""),
        "is_adhoc": slot.get("is_adhoc", False),
    }


def _get_active_window(slot_id):
    return mongo.db.voting_windows.find_one({"slot_id": slot_id, "is_active": True})


def _window_info(window):
    if not window:
        return {"id": None, "is_open": False, "opens_at": None, "closes_at": None}
    return {
        "id": str(window["_id"]),
        "is_open": is_voting_window_open(window["opens_at"], window["closes_at"]),
        "opens_at": format_ist(window["opens_at"]),
        "closes_at": format_ist(window["closes_at"]),
    }


# ── Dashboard ──────────────────────────────────────────────────────────────────

@admin_bp.route("/dashboard", methods=["GET"])
@admin_required
def dashboard():
    voters = list(mongo.db.users.find({"is_active": True, **VOTER_FILTER}).sort("name", 1))
    slots = list(mongo.db.match_slots.find({"is_active": {"$ne": False}}).sort("slot_number", 1))

    # Resolve each slot's own active window up front
    slot_windows = {str(s["_id"]): _get_active_window(str(s["_id"])) for s in slots}

    all_votes = list(mongo.db.votes.find({
        "$or": [
            {"slot_id": sid, "window_id": str(w["_id"])}
            for sid, w in slot_windows.items() if w
        ]
    })) if any(slot_windows.values()) else []
    vote_map = {(v["captain_id"], v["slot_id"]): v for v in all_votes}

    # Build voter × slot matrix
    matrix = []
    for voter in voters:
        cid = str(voter["_id"])
        row = {"captain": _user_to_dict(voter), "votes": []}
        for slot in slots:
            sid = str(slot["_id"])
            vote = vote_map.get((cid, sid))
            row["votes"].append({
                "slot_id": sid,
                "slot_label": f"Slot {slot['slot_number']}",
                "day": slot["day"],
                "time_of_day": slot["time_of_day"],
                "availability": vote["availability"] if vote else None,
                "voted_at": format_ist(vote["voted_at"]) if vote else None,
            })
        matrix.append(row)

    # Per-slot summary
    slot_summary = []
    open_count = 0
    for slot in slots:
        sid = str(slot["_id"])
        slot_votes = [v for v in all_votes if v["slot_id"] == sid]
        window_info = _window_info(slot_windows[sid])
        if window_info["is_open"]:
            open_count += 1
        slot_summary.append({
            "slot_id": sid,
            "slot_number": slot["slot_number"],
            "day": slot["day"],
            "time_of_day": slot["time_of_day"],
            "match_date": slot.get("match_date"),
            "is_adhoc": slot.get("is_adhoc", False),
            "label": f"Slot {slot['slot_number']} — {slot['day']} {slot['time_of_day']}",
            "available": sum(1 for v in slot_votes if v["availability"] == "available"),
            "not_available": sum(1 for v in slot_votes if v["availability"] == "not_available"),
            "maybe": sum(1 for v in slot_votes if v["availability"] == "maybe"),
            "no_response": len(voters) - len(slot_votes),
            "window": window_info,
            "weather": get_forecast_for_slot(slot),
        })

    return jsonify({
        "open_count": open_count,
        "total_slots": len(slots),
        "captains_total": len(voters),
        "captains_voted": len({v["captain_id"] for v in all_votes}),
        "slots": slot_summary,
        "vote_matrix": matrix,
    })


# ── Captains management ────────────────────────────────────────────────────────

@admin_bp.route("/captains", methods=["GET"])
@admin_required
def list_captains():
    captains = list(mongo.db.users.find({"role": "captain", "is_active": True}).sort("name", 1))
    return jsonify([_user_to_dict(c) for c in captains])


@admin_bp.route("/captains", methods=["POST"])
@admin_required
def add_captain():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    team_code = (data.get("team_code") or "").strip().upper()
    password = data.get("password") or team_code.lower()

    if not name or not team_code:
        return jsonify({"error": "name and team_code are required"}), 400

    if mongo.db.users.find_one({"team_code": team_code}):
        return jsonify({"error": "Team code already exists"}), 409

    doc = {
        "name": name,
        "team_code": team_code,
        "password_hash": generate_password_hash(password),
        "role": "captain",
        "is_active": True,
        "matches_scheduled": 0,
        "matches_played": 0,
        "tournament_status": "not_played",
        "must_change_password": True,
        "created_at": datetime.utcnow(),
    }
    result = mongo.db.users.insert_one(doc)
    doc["_id"] = result.inserted_id
    return jsonify({
        "message": "Captain added",
        "captain": _user_to_dict(doc),
        "default_password": password,
    }), 201


@admin_bp.route("/captains/<captain_id>", methods=["PUT"])
@admin_required
def update_captain(captain_id):
    data = request.get_json(silent=True) or {}
    updates = {}
    password_changed = False
    if "name" in data:
        updates["name"] = data["name"].strip()
    if "team_name" in data:
        updates["team_name"] = data["team_name"].strip()
    if "is_active" in data:
        updates["is_active"] = bool(data["is_active"])
    if "password" in data and data["password"]:
        password_error = validate_password(data["password"])
        if password_error:
            return jsonify({"error": password_error}), 400
        updates["password_hash"] = generate_password_hash(data["password"])
        updates["must_change_password"] = True
        password_changed = True
    if "team_code" in data:
        new_code = data["team_code"].strip().upper()
        if mongo.db.users.find_one({"team_code": new_code, "_id": {"$ne": ObjectId(captain_id)}}):
            return jsonify({"error": "Team code already taken"}), 409
        updates["team_code"] = new_code
    if "batting_average" in data:
        try:
            updates["batting_average"] = _parse_average(data["batting_average"])
        except ValueError as e:
            return jsonify({"error": f"batting_average {e}"}), 400
    if "bowling_average" in data:
        try:
            updates["bowling_average"] = _parse_average(data["bowling_average"])
        except ValueError as e:
            return jsonify({"error": f"bowling_average {e}"}), 400
    if "matches_scheduled" in data:
        updates["matches_scheduled"] = max(0, int(data["matches_scheduled"]))
    if "matches_played" in data:
        updates["matches_played"] = max(0, int(data["matches_played"]))
    if "tournament_status" in data:
        valid = {"not_played", "in_progress", "qualified", "eliminated"}
        if data["tournament_status"] not in valid:
            return jsonify({"error": "Invalid tournament_status"}), 400
        updates["tournament_status"] = data["tournament_status"]
    if "auction_category" in data:
        if data["auction_category"] not in AUCTION_CATEGORIES:
            return jsonify({"error": f"auction_category must be one of {sorted(AUCTION_CATEGORIES)}"}), 400
        updates["auction_category"] = data["auction_category"]
    if "role" in data:
        if data["role"] not in ("captain", "player"):
            return jsonify({"error": "role must be 'captain' or 'player'"}), 400
        updates["role"] = data["role"]

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    mongo_update = {"$set": updates}
    if password_changed:
        # Same session-revocation mechanism as self-service change-password —
        # an admin-set password must kick out whatever session this captain
        # was already using just as thoroughly as changing it themselves would.
        mongo_update["$inc"] = {"token_version": 1}
    result = mongo.db.users.update_one({"_id": ObjectId(captain_id)}, mongo_update)
    if result.matched_count == 0:
        return jsonify({"error": "Captain not found"}), 404

    updated = mongo.db.users.find_one({"_id": ObjectId(captain_id)})
    return jsonify({"message": "Captain updated", "captain": _user_to_dict(updated)})


@admin_bp.route("/captains/<captain_id>", methods=["DELETE"])
@admin_required
def remove_captain(captain_id):
    result = mongo.db.users.update_one(
        {"_id": ObjectId(captain_id), "role": "captain"},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        return jsonify({"error": "Captain not found"}), 404
    return jsonify({"message": "Captain deactivated"})


@admin_bp.route("/captains/<captain_id>/reset-device", methods=["POST"])
@admin_required
def reset_captain_device(captain_id):
    # Clears the bound device so the captain's next login registers whatever
    # phone/browser they use then — for a lost/replaced phone, not a way
    # around the lock for anyone but the account owner going forward.
    result = mongo.db.users.update_one(
        {"_id": ObjectId(captain_id), "role": "captain"},
        {"$unset": {"device_id": ""}}
    )
    if result.matched_count == 0:
        return jsonify({"error": "Captain not found"}), 404
    return jsonify({"message": "Device reset — next login will register a new device"})


@admin_bp.route("/captains/<captain_id>/reset-password", methods=["POST"])
@admin_required
def reset_captain_password(captain_id):
    return _reset_password(captain_id, "captain")


# ── Players management ──────────────────────────────────────────────────────────

@admin_bp.route("/players", methods=["GET"])
@admin_required
def list_players():
    # Captains are players too — they need the chance to vote even on a
    # weekend their own team has no match scheduled — so this roster is
    # everyone who can cast an availability vote, not just the dedicated
    # player-only accounts.
    players = list(mongo.db.users.find(
        {"is_active": True, **VOTER_FILTER}
    ).sort("name", 1))
    return jsonify([_user_to_dict(p) for p in players])


# ── Knockout attendance (reference-only — nothing else in the app reads or
# enforces these fields; admin tracks them to pick knockout lineups once
# league matches are done). attendance_count and total_matches_organized are
# both *derived* from league_matches (below) — admin records each match and
# checks off who actually showed up, rather than retyping running totals by
# hand for 57 people after every match. knockout_eligible is the one field
# that stays a manual, admin-editable override (see "Auto-Mark Top N" on the
# frontend, which pre-fills it from the computed ranking but never silently
# re-overwrites a hand-adjusted value on a later load). ─────────────────────

ATTENDANCE_SETTINGS_ID = "attendance"


def _attendance_settings():
    doc = mongo.db.settings.find_one({"_id": ATTENDANCE_SETTINGS_ID})
    return {
        "total_matches_organized": mongo.db.league_matches.count_documents({}),
        "knockout_cutoff": doc.get("knockout_cutoff", 28) if doc else 28,
    }


def _attendance_counts():
    """{voter_id_str: number of recorded matches they attended}."""
    counts = {}
    for match in mongo.db.league_matches.find({}, {"attendee_ids": 1}):
        for voter_id in match.get("attendee_ids", []):
            counts[voter_id] = counts.get(voter_id, 0) + 1
    return counts


@admin_bp.route("/attendance", methods=["GET"])
@admin_required
def list_attendance():
    voters = list(mongo.db.users.find(
        {"is_active": True, **VOTER_FILTER}
    ).sort("name", 1))
    counts = _attendance_counts()

    voter_dicts = []
    for v in voters:
        d = _user_to_dict(v)
        d["attendance_count"] = counts.get(str(v["_id"]), 0)
        voter_dicts.append(d)

    return jsonify({
        "voters": voter_dicts,
        "settings": _attendance_settings(),
    })


@admin_bp.route("/attendance/settings", methods=["PUT"])
@admin_required
def update_attendance_settings():
    data = request.get_json(silent=True) or {}

    cutoff = data.get("knockout_cutoff")
    if not isinstance(cutoff, int) or isinstance(cutoff, bool) or cutoff < 0:
        return jsonify({"error": "knockout_cutoff must be a non-negative integer"}), 400

    mongo.db.settings.update_one(
        {"_id": ATTENDANCE_SETTINGS_ID},
        {"$set": {"knockout_cutoff": cutoff}},
        upsert=True,
    )
    return jsonify({"message": "Settings updated", "settings": _attendance_settings()})


@admin_bp.route("/attendance", methods=["PUT"])
@admin_required
def update_attendance():
    data = request.get_json(silent=True) or {}
    updates = data.get("updates")
    if not isinstance(updates, list) or not updates:
        return jsonify({"error": "updates must be a non-empty list"}), 400

    ops = []
    for entry in updates:
        user_id = entry.get("id")
        if not user_id or not ObjectId.is_valid(user_id):
            return jsonify({"error": f"Invalid id: {user_id!r}"}), 400

        voter = mongo.db.users.find_one({"_id": ObjectId(user_id), "is_active": True, **VOTER_FILTER})
        if not voter:
            return jsonify({"error": f"{user_id} is not part of the voter roster"}), 400

        knockout_eligible = entry.get("knockout_eligible")
        if not isinstance(knockout_eligible, bool):
            return jsonify({"error": f"{voter['name']}: knockout_eligible must be true or false"}), 400

        ops.append(UpdateOne(
            {"_id": voter["_id"]},
            {"$set": {"knockout_eligible": knockout_eligible}},
        ))

    result = mongo.db.users.bulk_write(ops)
    return jsonify({"message": "Attendance updated", "modified_count": result.modified_count})


# ── League matches — one document per completed match, admin checks off who
# actually attended; attendance_count/total_matches_organized above are both
# derived from these. ───────────────────────────────────────────────────────

def _match_to_dict(m):
    return {
        "id": str(m["_id"]),
        "label": m.get("label", ""),
        "match_date": m.get("match_date"),
        "attendee_ids": m.get("attendee_ids", []),
        "attendee_count": len(m.get("attendee_ids", [])),
        "created_at": format_ist(m.get("created_at")),
    }


@admin_bp.route("/attendance/matches", methods=["GET"])
@admin_required
def list_league_matches():
    matches = list(mongo.db.league_matches.find({}).sort("created_at", 1))
    return jsonify([_match_to_dict(m) for m in matches])


@admin_bp.route("/attendance/matches", methods=["POST"])
@admin_required
def add_league_match():
    data = request.get_json(silent=True) or {}
    match_date = (data.get("match_date") or "").strip() or None

    label = (data.get("label") or "").strip()
    if not label:
        existing = mongo.db.league_matches.count_documents({})
        label = f"Match {existing + 1}"

    doc = {
        "label": label,
        "match_date": match_date,
        "attendee_ids": [],
        "created_at": datetime.utcnow(),
    }
    result = mongo.db.league_matches.insert_one(doc)
    doc["_id"] = result.inserted_id
    return jsonify({"message": "Match added", "match": _match_to_dict(doc)}), 201


@admin_bp.route("/attendance/matches/<match_id>", methods=["PUT"])
@admin_required
def update_league_match(match_id):
    if not ObjectId.is_valid(match_id):
        return jsonify({"error": "Invalid match id"}), 400
    match = mongo.db.league_matches.find_one({"_id": ObjectId(match_id)})
    if not match:
        return jsonify({"error": "Match not found"}), 404

    data = request.get_json(silent=True) or {}
    attendee_ids = data.get("attendee_ids")
    if not isinstance(attendee_ids, list):
        return jsonify({"error": "attendee_ids must be a list"}), 400

    for voter_id in attendee_ids:
        if not ObjectId.is_valid(voter_id):
            return jsonify({"error": f"Invalid id: {voter_id!r}"}), 400
        if not mongo.db.users.find_one({"_id": ObjectId(voter_id), "is_active": True, **VOTER_FILTER}):
            return jsonify({"error": f"{voter_id} is not part of the voter roster"}), 400

    mongo.db.league_matches.update_one(
        {"_id": match["_id"]}, {"$set": {"attendee_ids": attendee_ids}}
    )
    updated = mongo.db.league_matches.find_one({"_id": match["_id"]})
    return jsonify({"message": "Match attendance updated", "match": _match_to_dict(updated)})


@admin_bp.route("/attendance/matches/<match_id>", methods=["DELETE"])
@admin_required
def remove_league_match(match_id):
    if not ObjectId.is_valid(match_id):
        return jsonify({"error": "Invalid match id"}), 400
    result = mongo.db.league_matches.delete_one({"_id": ObjectId(match_id)})
    if result.deleted_count == 0:
        return jsonify({"error": "Match not found"}), 404
    return jsonify({"message": "Match removed"})


@admin_bp.route("/players", methods=["POST"])
@admin_required
def add_player():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    team_code = (data.get("team_code") or "").strip().upper()
    password = data.get("password") or team_code.lower()

    if not name or not team_code:
        return jsonify({"error": "name and team_code are required"}), 400

    if mongo.db.users.find_one({"team_code": team_code}):
        return jsonify({"error": "Player code already exists"}), 409

    doc = {
        "name": name,
        "team_code": team_code,
        "password_hash": generate_password_hash(password),
        "role": "player",
        "is_player": True,
        "is_active": True,
        "must_change_password": True,
        "created_at": datetime.utcnow(),
    }
    result = mongo.db.users.insert_one(doc)
    doc["_id"] = result.inserted_id
    return jsonify({
        "message": "Player added",
        "player": _user_to_dict(doc),
        "default_password": password,
    }), 201


@admin_bp.route("/players/<player_id>", methods=["PUT"])
@admin_required
def update_player(player_id):
    data = request.get_json(silent=True) or {}
    updates = {}
    password_changed = False
    if "name" in data:
        updates["name"] = data["name"].strip()
    if "is_active" in data:
        updates["is_active"] = bool(data["is_active"])
    if "password" in data and data["password"]:
        password_error = validate_password(data["password"])
        if password_error:
            return jsonify({"error": password_error}), 400
        updates["password_hash"] = generate_password_hash(data["password"])
        updates["must_change_password"] = True
        password_changed = True
    if "team_code" in data:
        new_code = data["team_code"].strip().upper()
        if mongo.db.users.find_one({"team_code": new_code, "_id": {"$ne": ObjectId(player_id)}}):
            return jsonify({"error": "Player code already taken"}), 409
        updates["team_code"] = new_code
    if "batting_average" in data:
        try:
            updates["batting_average"] = _parse_average(data["batting_average"])
        except ValueError as e:
            return jsonify({"error": f"batting_average {e}"}), 400
    if "bowling_average" in data:
        try:
            updates["bowling_average"] = _parse_average(data["bowling_average"])
        except ValueError as e:
            return jsonify({"error": f"bowling_average {e}"}), 400
    if "auction_category" in data:
        if data["auction_category"] not in AUCTION_CATEGORIES:
            return jsonify({"error": f"auction_category must be one of {sorted(AUCTION_CATEGORIES)}"}), 400
        updates["auction_category"] = data["auction_category"]
    if "team_name" in data:
        updates["team_name"] = data["team_name"].strip()
    if "role" in data:
        # Promotes an existing player to captain (or the reverse) in place —
        # keeps their real login (team_code/password) untouched, unlike creating
        # a brand-new captain record under a different code.
        if data["role"] not in ("captain", "player"):
            return jsonify({"error": "role must be 'captain' or 'player'"}), 400
        updates["role"] = data["role"]

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    mongo_update = {"$set": updates}
    if password_changed:
        mongo_update["$inc"] = {"token_version": 1}
    result = mongo.db.users.update_one(
        {"_id": ObjectId(player_id), "$or": [{"role": "player"}, ADMIN_VOTER_FILTER]},
        mongo_update,
    )
    if result.matched_count == 0:
        return jsonify({"error": "Player not found"}), 404

    updated = mongo.db.users.find_one({"_id": ObjectId(player_id)})
    return jsonify({"message": "Player updated", "player": _user_to_dict(updated)})


@admin_bp.route("/players/<player_id>", methods=["DELETE"])
@admin_required
def remove_player(player_id):
    result = mongo.db.users.update_one(
        {"_id": ObjectId(player_id), "role": "player"},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        return jsonify({"error": "Player not found"}), 404
    return jsonify({"message": "Player deactivated"})


@admin_bp.route("/players/<player_id>/reset-device", methods=["POST"])
@admin_required
def reset_player_device(player_id):
    result = mongo.db.users.update_one(
        {"_id": ObjectId(player_id), "role": "player"},
        {"$unset": {"device_id": ""}}
    )
    if result.matched_count == 0:
        return jsonify({"error": "Player not found"}), 404
    return jsonify({"message": "Device reset — next login will register a new device"})


@admin_bp.route("/players/<player_id>/reset-password", methods=["POST"])
@admin_required
def reset_player_password(player_id):
    return _reset_password(player_id, "player")


# ── Ad-hoc match slots ───────────────────────────────────────────────────────────
# The 4 recurring weekend slots are seeded once (backend/scripts/seed.py) and never
# created at runtime. These two endpoints let admin add a one-off dated slot (e.g.
# a public holiday or a weather-driven date) on top of those — everything else
# (windows, voting, summaries, exports) already queries match_slots generically,
# so a new slot just shows up everywhere those already iterate the collection.

@admin_bp.route("/slots", methods=["POST"])
@admin_required
def add_slot():
    data = request.get_json(silent=True) or {}
    match_date = (data.get("match_date") or "").strip()
    day = (data.get("day") or "").strip()
    time_of_day = (data.get("time_of_day") or "").strip()
    description = (data.get("description") or "").strip()

    if not match_date or not day or not time_of_day:
        return jsonify({"error": "match_date, day and time_of_day are required"}), 400

    try:
        datetime.fromisoformat(match_date)
    except ValueError:
        return jsonify({"error": "match_date must be an ISO date, e.g. 2026-08-15"}), 400

    last_slot = mongo.db.match_slots.find_one(sort=[("slot_number", -1)])
    next_number = (last_slot["slot_number"] + 1) if last_slot else 1

    doc = {
        "slot_number": next_number,
        "day": day,
        "time_of_day": time_of_day,
        # SlotCard.jsx already prefers match_time over time_of_day for its bold
        # headline — reuse that unchanged by feeding the admin's description in
        # here, so e.g. "Independence Day Match" displays instead of just "Morning".
        "match_time": description or time_of_day,
        "description": description,
        "match_date": match_date,
        "is_adhoc": True,
        "is_active": True,
        "created_at": datetime.utcnow(),
    }
    result = mongo.db.match_slots.insert_one(doc)
    doc["_id"] = result.inserted_id
    return jsonify({"message": "Ad-hoc match added", "slot": _slot_to_dict(doc)}), 201


@admin_bp.route("/slots/<slot_id>", methods=["DELETE"])
@admin_required
def remove_slot(slot_id):
    result = mongo.db.match_slots.update_one(
        {"_id": ObjectId(slot_id), "is_adhoc": True},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        return jsonify({"error": "Ad-hoc match not found (only ad-hoc matches can be removed)"}), 404
    return jsonify({"message": "Ad-hoc match removed"})


# ── Voting window management (per match slot) ──────────────────────────────────

@admin_bp.route("/window", methods=["GET"])
@admin_required
def get_window():
    slots = list(mongo.db.match_slots.find({"is_active": {"$ne": False}}).sort("slot_number", 1))
    windows = []
    for slot in slots:
        sid = str(slot["_id"])
        window = _get_active_window(sid)
        suggested = suggested_window_for_slot(slot)
        slot_dict = _slot_to_dict(slot)
        slot_dict["weather"] = get_forecast_for_slot(slot)
        windows.append({
            "slot": slot_dict,
            "window": _window_info(window) if window else None,
            "suggested": {
                "opens_at": suggested["opens_at_display"],
                "closes_at": suggested["closes_at_display"],
                "opens_at_iso": suggested["opens_at_ist_iso"],
                "closes_at_iso": suggested["closes_at_ist_iso"],
            } if suggested else None,
        })
    return jsonify({"windows": windows})


@admin_bp.route("/window", methods=["POST"])
@admin_required
def set_window():
    """
    Body: { "slot_id": "...", "opens_at": "2024-06-06T12:30:00", "closes_at": "2024-06-07T14:30:00" }
    Datetimes treated as IST.
    """
    data = request.get_json(silent=True) or {}
    slot_id = data.get("slot_id")
    if not slot_id or not mongo.db.match_slots.find_one({"_id": ObjectId(slot_id)}):
        return jsonify({"error": "Valid slot_id is required"}), 400

    try:
        opens_str = data["opens_at"]
        closes_str = data["closes_at"]
        opens_naive = datetime.fromisoformat(opens_str)
        closes_naive = datetime.fromisoformat(closes_str)
        opens_at = IST.localize(opens_naive).astimezone(pytz.utc).replace(tzinfo=None)
        closes_at = IST.localize(closes_naive).astimezone(pytz.utc).replace(tzinfo=None)
    except (KeyError, ValueError) as e:
        return jsonify({"error": f"Invalid datetime: {e}"}), 400

    if opens_at >= closes_at:
        return jsonify({"error": "opens_at must be before closes_at"}), 400

    # Deactivate previous windows for this slot only
    mongo.db.voting_windows.update_many({"slot_id": slot_id}, {"$set": {"is_active": False}})
    result = mongo.db.voting_windows.insert_one({
        "slot_id": slot_id,
        "opens_at": opens_at,
        "closes_at": closes_at,
        "is_active": True,
        "created_at": datetime.utcnow(),
    })
    return jsonify({
        "message": "Voting window set",
        "window_id": str(result.inserted_id),
        "opens_at": format_ist(opens_at),
        "closes_at": format_ist(closes_at),
    }), 201


@admin_bp.route("/window/close", methods=["POST"])
@admin_required
def close_window_early():
    data = request.get_json(silent=True) or {}
    slot_id = data.get("slot_id")
    if not slot_id:
        return jsonify({"error": "slot_id is required"}), 400

    window = _get_active_window(slot_id)
    if not window:
        return jsonify({"error": "No active window to close for this slot"}), 404
    now = datetime.utcnow()
    mongo.db.voting_windows.update_one(
        {"_id": window["_id"]},
        {"$set": {"closes_at": now, "closed_early": True}}
    )
    return jsonify({"message": "Voting window closed early"})


# ── Export ─────────────────────────────────────────────────────────────────────

def _votes_for_current_windows(slots):
    slot_windows = {str(s["_id"]): _get_active_window(str(s["_id"])) for s in slots}
    if not any(slot_windows.values()):
        return []
    return list(mongo.db.votes.find({
        "$or": [
            {"slot_id": sid, "window_id": str(w["_id"])}
            for sid, w in slot_windows.items() if w
        ]
    }))


@admin_bp.route("/export/csv", methods=["GET"])
@admin_required
def export_csv():
    captains = list(mongo.db.users.find({"is_active": True, **VOTER_FILTER}).sort("name", 1))
    slots = list(mongo.db.match_slots.find({"is_active": {"$ne": False}}).sort("slot_number", 1))
    votes = _votes_for_current_windows(slots)

    csv_data = build_csv_report(captains, slots, votes)
    response = make_response(csv_data)
    response.headers["Content-Type"] = "text/csv"
    response.headers["Content-Disposition"] = (
        f"attachment; filename=bcc-cvote-availability-{datetime.utcnow().strftime('%Y%m%d')}.csv"
    )
    return response


@admin_bp.route("/export/excel", methods=["GET"])
@admin_required
def export_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    captains = list(mongo.db.users.find({"is_active": True, **VOTER_FILTER}).sort("name", 1))
    slots = list(mongo.db.match_slots.find({"is_active": {"$ne": False}}).sort("slot_number", 1))
    votes = _votes_for_current_windows(slots)
    vote_map = {(v["captain_id"], v["slot_id"]): v["availability"] for v in votes}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Availability"

    # Fills
    green  = PatternFill("solid", fgColor="C6EFCE")
    red    = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    grey   = PatternFill("solid", fgColor="EFEFEF")
    navy   = PatternFill("solid", fgColor="1E3A5F")

    bold_white = Font(bold=True, color="FFFFFF")
    bold       = Font(bold=True)
    center     = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Header row
    headers = ["#", "Captain", "Team Name", "Team Code"] + [
        f"{s['day']}\n{s.get('match_time', s['time_of_day'])}" for s in slots
    ] + ["Total Available"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = navy
        cell.font = bold_white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 36

    # Data rows
    avail_map = {
        "available":     ("✅ Available",     green),
        "not_available": ("❌ Not Available", red),
        "maybe":         ("🤔 Maybe",         yellow),
        None:            ("—",               grey),
    }

    for row_idx, captain in enumerate(captains, 2):
        cid = str(captain["_id"])
        avail_count = 0
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center
        ws.cell(row=row_idx, column=2, value=captain["name"]).font = bold
        ws.cell(row=row_idx, column=3, value=captain.get("team_name", ""))
        ws.cell(row=row_idx, column=4, value=captain["team_code"]).alignment = center

        for col_idx, slot in enumerate(slots, 5):
            sid = str(slot["_id"])
            avail = vote_map.get((cid, sid))
            label, fill = avail_map.get(avail, avail_map[None])
            cell = ws.cell(row=row_idx, column=col_idx, value=label)
            cell.fill = fill
            cell.alignment = center
            cell.border = thin
            if avail == "available":
                avail_count += 1

        total_cell = ws.cell(row=row_idx, column=5 + len(slots), value=avail_count)
        total_cell.alignment = center
        total_cell.font = bold
        total_cell.border = thin

        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = thin

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 10
    for i in range(len(slots)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 18
    ws.column_dimensions[get_column_letter(5 + len(slots))].width = 16

    # Summary row
    sum_row = len(captains) + 2
    ws.cell(row=sum_row, column=1, value="").fill = grey
    ws.cell(row=sum_row, column=2, value="AVAILABLE COUNT").font = bold
    ws.cell(row=sum_row, column=3, value="").fill = grey
    ws.cell(row=sum_row, column=4, value="").fill = grey
    for col_idx, slot in enumerate(slots, 5):
        sid = str(slot["_id"])
        count = sum(1 for v in votes if v["slot_id"] == sid and v["availability"] == "available")
        cell = ws.cell(row=sum_row, column=col_idx, value=count)
        cell.font = bold
        cell.alignment = center
        cell.fill = green
        cell.border = thin

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"BCC-Availability-{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@admin_bp.route("/export/available-players", methods=["GET"])
@admin_required
def export_available_players():
    """
    One sheet per match slot, listing just the players who voted "available" —
    a ready-to-use roster for picking teams/lineups, instead of the full
    everyone-x-every-status grid in /export/excel.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    voters = {str(u["_id"]): u for u in mongo.db.users.find(
        {"is_active": True, **VOTER_FILTER}
    )}
    slots = list(mongo.db.match_slots.find({"is_active": {"$ne": False}}).sort("slot_number", 1))
    votes = _votes_for_current_windows(slots)

    navy = PatternFill("solid", fgColor="1E3A5F")
    gold = PatternFill("solid", fgColor="FFF3CD")
    bold_white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(*(Side(style="thin"),) * 4)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for slot in slots:
        sid = str(slot["_id"])
        sheet_name = f"{slot['day']} {slot['time_of_day']}"[:31]
        ws = wb.create_sheet(sheet_name)

        title = f"{slot['day']} {slot['time_of_day']} Match — {slot.get('match_time', '')}"
        ws.merge_cells("A1:C1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13, color="1E3A5F")

        headers = ["#", "Player Name", "Code"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = navy
            cell.font = bold_white
            cell.alignment = center
            cell.border = thin

        available = sorted(
            (voters[v["captain_id"]] for v in votes
             if v["slot_id"] == sid and v["availability"] == "available" and v["captain_id"] in voters),
            key=lambda u: u["name"].lower(),
        )

        row = 4
        for i, user in enumerate(available, 1):
            ws.cell(row=row, column=1, value=i).alignment = center
            ws.cell(row=row, column=2, value=user["name"])
            ws.cell(row=row, column=3, value=user["team_code"]).alignment = center
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin
            row += 1

        summary_row = max(row, 5)
        ws.cell(row=summary_row + 1, column=1, value="Available:").font = bold
        ws.cell(row=summary_row + 1, column=2, value=len(available)).fill = gold

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 12

        if not available:
            ws.cell(row=4, column=1, value="No one has voted available for this match yet.")
            ws.merge_cells("A4:C4")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"BCC-Available-Players-{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)
